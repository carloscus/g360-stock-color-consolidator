"""G360 Stock Color Consolidator - Aplicacion principal."""
from __future__ import annotations

import json
import os
import shutil
import tempfile
from datetime import datetime
from pathlib import Path

import flet as ft
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Core imports
from src.config.theme import LIGHT, DARK, Modo
from src.core.parsers import parse_source2, agregar_almacenes
from src.core.consolidator import consolidar
from src.core.downloader import download_source1
from src.core.browser_automation import download_source2 as browser_download_source2
from src.core.models import ProductoConsolidado

# UI imports
from src.ui.dashboard import Dashboard
from src.ui.sku_detail import SkuDetailModal

CREDENTIALS_DIR = Path(os.environ.get("APPDATA", Path.home())) / "g360-stock-consolidator"
CREDENTIALS_FILE = CREDENTIALS_DIR / "creds.json"


def _generar_reporte_xlsx(
    productos: list[ProductoConsolidado],
    path: str,
    source1_raw: dict[str, dict[str, dict]] | None = None,
):
    wb = Workbook()

    warehouses = sorted(
        c for c in (source1_raw or {}).keys() if c != "VES"
    )

    username = os.environ.get("G360_S2_USER", "")
    today_str = datetime.now().strftime("%d/%m/%Y")

    def _warehouse_vals(sku: str) -> list:
        if not source1_raw:
            return [None] * len(warehouses)
        return [
            max(0, source1_raw.get(w, {}).get(sku, {}).get("stock", 0)
                - source1_raw.get(w, {}).get(sku, {}).get("predespacho", 0))
            or None
            for w in warehouses
        ]

    def _has_real_color(p: ProductoConsolidado) -> bool:
        return any(c.nombre != "SIN COLOR" for c in p.colores)

    # ── Styles ─────────────────────────────────────────────────────────
    title_font = Font(bold=True, size=14, color="ffffff")
    title_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_font = Font(bold=True, size=11, color="ffffff")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def _write_sheet(ws, title_suffix, headers, data_rows):
        ncols = len(headers)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(row=1, column=1)
        c.value = f"Reporte de Stock - Colores | {title_suffix} | Usuario: {username} | {today_str}"
        c.font = title_font
        c.fill = title_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for ri, row_data in enumerate(data_rows, start=3):
            for ci, val in enumerate(row_data, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = thin_border

        ws.freeze_panes = "A3"

        for col in ws.columns:
            real_cells = [c for c in col if not isinstance(c, MergedCell)]
            if not real_cells:
                continue
            max_len = max((len(str(c.value or "")) for c in real_cells), default=8)
            ws.column_dimensions[real_cells[0].column_letter].width = min(max_len + 3, 40)

    # ── Sheet 1: Con Color ─────────────────────────────────────────────
    headers1 = ["#", "SKU", "Descripción", "Stock", "Predesp", "Disponible",
                 "Modelo", "Color", "Cantidad"]
    for w in warehouses:
        headers1.append(w)

    rows1 = []
    i = 1
    for p in sorted(productos, key=lambda x: x.sku):
        if not _has_real_color(p):
            continue
        wh_vals = _warehouse_vals(p.sku)
        for c in sorted(p.colores, key=lambda x: x.nombre):
            if c.nombre == "SIN COLOR":
                continue
            if c.disenos:
                for d in sorted(c.disenos, key=lambda x: x.nombre):
                    rows1.append([i, p.sku, p.descripcion, p.stock_referencial,
                                  p.predespacho_total, p.disponible,
                                  d.nombre, c.nombre, d.cantidad] + wh_vals)
                    i += 1
            else:
                rows1.append([i, p.sku, p.descripcion, p.stock_referencial,
                              p.predespacho_total, p.disponible,
                              "", c.nombre, c.total] + wh_vals)
                i += 1

    ws1 = wb.active
    ws1.title = "Con Color"
    _write_sheet(ws1, "Con Color", headers1, rows1)

    # ── Sheet 2: Sin Color ─────────────────────────────────────────────
    headers2 = ["#", "SKU", "Descripción", "Stock", "Predesp", "Disponible"]
    for w in warehouses:
        headers2.append(w)

    rows2 = []
    for j, p in enumerate(sorted(productos, key=lambda x: x.sku), start=1):
        if _has_real_color(p):
            continue
        rows2.append([j, p.sku, p.descripcion, p.stock_referencial,
                       p.predespacho_total, p.disponible] + _warehouse_vals(p.sku))

    ws2 = wb.create_sheet("Sin Color")
    _write_sheet(ws2, "Sin Color", headers2, rows2)

    wb.save(path)


def _friendly_error(ex: Exception) -> str:
    msg = str(ex)
    exc_type = type(ex).__name__
    if exc_type == "TimeoutError" or "Timeout" in exc_type or "timed out" in msg.lower():
        return "La conexión con el servidor no respondió a tiempo. Verifique su conexión a internet e intente nuevamente."
    if "Credenciales incorrectas" in msg or "credenciales" in msg.lower():
        return "Credenciales incorrectas o sin acceso al ERP. Verifique usuario y contraseña en el diálogo de credenciales."
    if msg == "" or ("user" in msg.lower() and "pass" in msg.lower()):
        return "Debe ingresar usuario y contraseña para descargar del ERP."
    if exc_type in ("ConnectionError", "ConnectionRefusedError", "ConnectionAbortedError", "ConnectionResetError"):
        return "No se pudo conectar al servidor. Verifique su conexión a internet o que el servidor esté disponible."
    if "ENOTFOUND" in msg or "getaddrinfo" in msg:
        return "No se pudo resolver la dirección del servidor. Verifique su conexión a internet."
    if "ECONNREFUSED" in msg:
        return "El servidor rechazó la conexión. Puede estar caído o no accesible desde su red."
    if "HTTPError" in exc_type or "status" in msg.lower():
        import re
        match = re.search(r"(\d{3})", msg)
        code = match.group(1) if match else ""
        return f"El servidor respondió con un error (HTTP {code}). Intente más tarde." if code else "El servidor respondió con un error inesperado."
    if "Playwright" in msg or "playwright" in msg.lower():
        return "Error al controlar el navegador automático. Revise los logs para más detalles."
    if "parse" in msg.lower() or "parsing" in msg.lower():
        return "Error al procesar el archivo descargado. El formato puede ser incorrecto."
    # fallback: short, clean message
    short = msg if len(msg) < 150 else msg[:147] + "..."
    return f"Error inesperado: {short}"


def _show_error(page: ft.Page, title: str, ex: Exception, close_cb=None, on_retry_creds=None, on_manual_file=None):
    message = _friendly_error(ex)
    is_cred = "credenciales" in message.lower()

    def _change(e):
        _close_dlg(page, dlg, None)
        if callable(on_retry_creds):
            on_retry_creds()

    def _manual(e):
        _close_dlg(page, dlg, None)
        if callable(on_manual_file):
            on_manual_file()

    actions = [
        ft.TextButton(
            "OK",
            on_click=lambda _: _close_dlg(page, dlg, close_cb),
        ),
    ]
    if is_cred and on_retry_creds:
        actions.insert(0, ft.ElevatedButton("Cambiar credenciales", on_click=_change))
    if on_manual_file:
        actions.insert(0, ft.TextButton("Cargar manualmente", on_click=_manual))

    dlg = ft.AlertDialog(
        title=ft.Text(title),
        content=ft.Text(message),
        actions=actions,
    )
    page.dialog = dlg
    dlg.open = True
    page.update()


def _validar_archivo_source2(path: str) -> bool:
    try:
        with open(path, "rb") as f:
            raw = f.read(2000)
        content = raw.decode("utf-8", errors="replace").lower()
        return "<html" in content and "color" in content and "cantidad" in content
    except Exception:
        return False


def _close_dlg(page: ft.Page, dlg: ft.AlertDialog, extra_cb=None):
    dlg.open = False
    page.update()
    if callable(extra_cb):
        extra_cb()


class StockConsolidatorApp:
    def __init__(self, page: ft.Page):
        self.page = page
        self.modo = Modo.LIGHT
        self.source1_raw: dict[str, dict[str, dict]] = {}
        self.source2_data: dict[str, list[tuple[str, str, int]]] = {}
        self.productos: list[ProductoConsolidado] = []
        self.dashboard: Dashboard | None = None
        self._picker: ft.FilePicker | None = None
        self._save_picker: ft.FilePicker | None = None
        self._creds: dict[str, str] = self._load_cached_credentials()
        self._setup_window()
        self._build()
        self._show_credentials_dialog()

    def _setup_window(self):
        self.page.title = "G360 - Stock Color Consolidator"
        self.page.theme_mode = ft.ThemeMode.LIGHT
        self.page.padding = 0
        self.page.window_width = 960
        self.page.window_height = 780
        self.page.window_min_width = 720
        self.page.window_min_height = 600
        self.page.window_center()
        self.page.bgcolor = LIGHT.bg

    def _build(self):
        paleta = DARK if self.modo == Modo.DARK else LIGHT
        self.page.bgcolor = paleta.bg

        self.dashboard = Dashboard(self.page)
        self.dashboard.set_on_sku_click(self._on_sku_click)
        self.dashboard.set_on_theme_toggle(self._on_theme_toggle)
        self.dashboard.set_on_download_report(self._on_download_report)
        self.dashboard.set_on_credentials(self._show_credentials_dialog)
        self.dashboard.set_on_expand_all(self._on_expand_all)
        self.dashboard.set_on_load_source(
            self._on_download_source1,
            self._on_download_source2,
            self._pick_file,
        )
        self.dashboard.set_source1_raw(self.source1_raw)
        self.dashboard.set_theme(self.modo)

        self._save_picker = ft.FilePicker(on_result=self._on_save_report)
        self.page.overlay.append(self._save_picker)

        view = self.dashboard.build()

        self.page.clean()
        self.page.add(view)
        
        # Cargamos productos
        self.dashboard.set_productos(self.productos)
        self.page.update()

    def _on_expand_all(self, expand: bool):
        if self.dashboard:
            if hasattr(self.dashboard, "expand_all"):
                self.dashboard.expand_all(expand)

    def _on_theme_toggle(self):
        self.modo = Modo.DARK if self.modo == Modo.LIGHT else Modo.LIGHT
        paleta = DARK if self.modo == Modo.DARK else LIGHT
        self.page.theme_mode = ft.ThemeMode.DARK if self.modo == Modo.DARK else ft.ThemeMode.LIGHT
        self.page.bgcolor = paleta.bg
        self._build()

    def _on_download_source1(self):
        if self.source1_raw:
            self._confirm_reload(
                "Ya hay datos de Source 1 cargados.\n"
                "Solo descargue de nuevo si hubo movimientos de stock nuevos.\n"
                "¿Desea continuar?",
                self._run_download_source1,
            )
            return
        self._run_download_source1()

    def _run_download_source1(self):
        self.dashboard.set_loading(True, "Descargando Source 1...")
        try:
            result = download_source1()
            if not result:
                self.dashboard.set_loading(False)
                self._show_toast("No se obtuvieron datos del servidor.")
                return
            self.source1_raw = result
            self.productos = consolidar(
                agregar_almacenes(self.source1_raw, ["VES"]),
                self.source2_data,
            )
            if self.dashboard:
                self.dashboard.set_source1_raw(self.source1_raw)
                self.dashboard.set_productos(self.productos)
            self.page.update()
            self._show_toast("Source 1 descargado correctamente.")
        except Exception as ex:
            import traceback
            traceback.print_exc()
            _show_error(self.page, "Error al descargar Source 1", ex)
        finally:
            self.dashboard.set_loading(False)
            self.page.update()

    def _on_download_source2(self):
        if self.source2_data:
            self._confirm_reload(
                "Ya hay datos de Source 2 procesados.\n"
                "Solo descargue de nuevo si hay nuevos colores o productos.\n"
                "¿Desea continuar?",
                self._run_download_source2_with_creds,
            )
            return
        self._run_download_source2_with_creds()

    def _run_download_source2_with_creds(self):
        if not self._creds.get("user") or not self._creds.get("pass"):
            self._show_credentials_dialog(on_save_callback=self._run_download_source2)
            return
        self._run_download_source2()

    def _run_download_source2(self):
        download_dir: str | None = None
        os.environ["G360_S2_USER"] = self._creds.get("user", "")
        os.environ["G360_S2_PASS"] = self._creds.get("pass", "")
        self.dashboard.set_loading(True, "Descargando Source 2 (ERP)...")
        try:
            path = browser_download_source2(
                progress_callback=lambda msg: self.dashboard.set_loading(True, msg)
            )
            if not path:
                self.dashboard.set_loading(False)
                self._show_toast("No se obtuvo archivo del servidor.")
                return

            download_dir = str(Path(path).parent)

            file_size = Path(path).stat().st_size
            if file_size < 100:
                self.dashboard.set_loading(False)
                self._show_toast(f"Archivo muy pequeno ({file_size} bytes). Revise logs.")
                return

            if not _validar_archivo_source2(path):
                self.dashboard.set_loading(False)
                self._show_toast("El archivo descargado no tiene el formato esperado del ERP.")
                return

            self.dashboard.set_loading(True, "Procesando colores...")
            self.source2_data = parse_source2(path)
            if not self.source2_data:
                self.dashboard.set_loading(False)
                self._show_toast("El archivo descargado no contiene datos de colores. Revise logs.")
                return
            self.productos = consolidar(
                agregar_almacenes(self.source1_raw, ["VES"]),
                self.source2_data,
            )
            if self.dashboard:
                self.dashboard.set_productos(self.productos)
            self.page.update()
            self._show_toast(f"Source 2 descargado: {Path(path).name} ({len(self.source2_data)} SKUs)")
        except Exception as ex:
            import traceback
            traceback.print_exc()
            _show_error(self.page, "Error al descargar Source 2", ex, on_retry_creds=self._show_credentials_dialog, on_manual_file=self._pick_file)
        finally:
            self.dashboard.set_loading(False)
            self._limpiar_descarga(download_dir)
            self.page.update()

    def _on_sku_click(self, sku: str):
        producto = next((p for p in self.productos if p.sku == sku), None)
        if producto:
            paleta = DARK if self.modo == Modo.DARK else LIGHT
            modal = SkuDetailModal(self.page, producto, paleta)
            modal.show()

    def _pick_file(self):
        if self._picker:
            self.page.overlay.remove(self._picker)
        self._picker = ft.FilePicker(on_result=self._on_file_picked)
        self.page.overlay.append(self._picker)
        self.page.update()

        self._picker.pick_files(
            allow_multiple=False,
            allowed_extensions=["xlsx", "xls", "csv"],
            dialog_title="Seleccionar Source 2 (Colores)",
        )

    def _on_file_picked(self, e: ft.FilePickerResultEvent):
        if not e.files or not e.files[0].path:
            return

        path = e.files[0].path
        self.dashboard.set_loading(True, "Procesando archivo de colores...")
        try:
            self.source2_data = parse_source2(path)
            self.productos = consolidar(
                agregar_almacenes(self.source1_raw, ["VES"]),
                self.source2_data,
            )

            if self.dashboard:
                self.dashboard.set_productos(self.productos)
            self.page.update()
        except Exception as ex:
            import traceback
            traceback.print_exc()
            _show_error(self.page, "Error al procesar archivo", ex)
        finally:
            self.dashboard.set_loading(False)
            self.page.update()

    def _close_dialog(self, dlg: ft.AlertDialog, extra_cb=None):
        _close_dlg(self.page, dlg, extra_cb)

    def _confirm_reload(self, message: str, on_confirm):
        dlg = ft.AlertDialog(
            title=ft.Text("Confirmar descarga"),
            content=ft.Text(message),
            actions=[
                ft.TextButton("Cancelar", on_click=lambda _: self._close_dialog(dlg)),
                ft.ElevatedButton("Descargar", on_click=lambda _: self._close_dialog(dlg, on_confirm)),
            ],
        )
        self.page.dialog = dlg
        dlg.open = True
        # Usamos update del diálogo específicamente si es posible, 
        # o aseguramos el update de la página.
        self.page.update()

    def _limpiar_descarga(self, download_dir: str | None):
        if download_dir and "g360_s2_" in download_dir:
            try:
                shutil.rmtree(download_dir, ignore_errors=True)
            except Exception:
                pass

    def _load_cached_credentials(self) -> dict[str, str]:
        if CREDENTIALS_FILE.exists():
            try:
                return json.loads(CREDENTIALS_FILE.read_text())
            except Exception:
                pass
        return {"user": "", "pass": ""}

    def _cache_credentials(self):
        CREDENTIALS_DIR.mkdir(parents=True, exist_ok=True)
        safe = {"user": self._creds.get("user", "")}
        CREDENTIALS_FILE.write_text(json.dumps(safe))

    def _show_credentials_dialog(self, on_save_callback=None):
        hint_user = self._creds.get("user", "")
        user_field = ft.TextField(
            label="Usuario",
            hint_text=hint_user if hint_user else "usuario",
            autofocus=True,
        )
        pass_field = ft.TextField(
            label="Contraseña",
            password=True,
            can_reveal_password=True,
        )

        def _on_save(e):
            self._creds["user"] = user_field.value.strip()
            self._creds["pass"] = pass_field.value
            self._cache_credentials()
            self._close_dialog(dlg)
            self._show_toast("Credenciales guardadas.")
            if callable(on_save_callback):
                on_save_callback()

        dlg = ft.AlertDialog(
            title=ft.Text("Credenciales ERP (appweb.cipsa.com.pe)"),
            content=ft.Column([user_field, pass_field], tight=True, width=320),
            actions=[
                ft.TextButton("Cancelar", on_click=lambda _: self._close_dialog(dlg)),
                ft.ElevatedButton("Guardar", on_click=_on_save),
            ],
        )
        self.page.dialog = dlg
        dlg.open = True
        self.page.update()

    def _on_download_report(self):
        if not self.productos:
            self._show_toast("No hay datos. Cargue Source 1 y Source 2 primero.")
            return
        self._save_picker.save_file(
            file_name=f"reporte_stock_colores_{datetime.now():%d%m%Y}.xlsx",
            allowed_extensions=["xlsx"],
            dialog_title="Guardar reporte XLSX",
        )

    def _on_save_report(self, e: ft.FilePickerResultEvent):
        if not e.path:
            return
        try:
            os.environ["G360_S2_USER"] = self._creds.get("user", "")
            _generar_reporte_xlsx(self.productos, e.path, self.source1_raw)
            self._show_toast(f"Reporte guardado: {e.path}")
        except Exception as ex:
            import traceback
            traceback.print_exc()
            _show_error(self.page, "Error al guardar reporte", ex)

    def _show_toast(self, msg: str):
        snack = ft.SnackBar(content=ft.Text(msg), duration=3000)
        self.page.overlay.append(snack)
        snack.open = True
        self.page.update()


def main(page: ft.Page):
    StockConsolidatorApp(page)
