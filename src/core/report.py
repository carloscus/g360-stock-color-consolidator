"""Generación de reportes XLSX."""
from __future__ import annotations

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from src.core.constants import (
    COLOR_SIN_COLOR,
    REPORT_FILENAME_FORMAT,
    REPORT_FILENAME_PREFIX,
)
from src.core.models import ProductoConsolidado


def generar_reporte_xlsx(
    productos: list[ProductoConsolidado],
    path: str,
    source1_raw: dict[str, dict[str, dict]] | None = None,
):
    """Genera un reporte XLSX con dos hojas: Con Color y Sin Color."""
    wb = Workbook()

    warehouses = sorted(
        c for c in (source1_raw or {}).keys() if c != "VES"
    )

    username = os.environ.get("G360_S2_USER", "")
    today_str = datetime.now().strftime(REPORT_FILENAME_FORMAT)

    def _warehouse_vals(sku: str) -> list:
        if not source1_raw:
            return [None] * len(warehouses)
        return [
            max(0, source1_raw.get(w, {}).get(sku, {}).get("stock", 0)
                - source1_raw.get(w, {}).get(sku, {}).get("predespacho", 0))
            or None
            for w in warehouses
        ]

    def _has_real_color(p: ProductoConsolidado) -> bool:
        return any(c.nombre != COLOR_SIN_COLOR for c in p.colores)

    # ── Styles ─────────────────────────────────────────────────────────
    title_font = Font(bold=True, size=14, color="ffffff")
    title_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_font = Font(bold=True, size=11, color="ffffff")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def _write_sheet(ws, title_suffix, headers, data_rows):
        ncols = len(headers)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(row=1, column=1)
        c.value = f"Reporte de Stock - Colores | {title_suffix} | Usuario: {username} | {today_str}"
        c.font = title_font
        c.fill = title_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for ri, row_data in enumerate(data_rows, start=3):
            for ci, val in enumerate(row_data, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = thin_border

        ws.freeze_panes = "A3"

        for col in ws.columns:
            real_cells = [c for c in col if not isinstance(c, MergedCell)]
            if not real_cells:
                continue
            max_len = max((len(str(c.value or "")) for c in real_cells), default=8)
            ws.column_dimensions[real_cells[0].column_letter].width = min(max_len + 3, 40)

    # ── Sheet 1: Con Color ─────────────────────────────────────────────
    headers1 = ["#", "SKU", "Descripción", "Stock", "Predesp", "Disponible",
                 "Modelo", "Color", "Cantidad"]
    for w in warehouses:
        headers1.append(w)

    rows1 = []
    i = 1
    for p in sorted(productos, key=lambda x: x.sku):
        if not _has_real_color(p):
            continue
        wh_vals = _warehouse_vals(p.sku)
        for c in sorted(p.colores, key=lambda x: x.nombre):
            if c.nombre == COLOR_SIN_COLOR:
                continue
            if c.disenos:
                for d in sorted(c.disenos, key=lambda x: x.nombre):
                    rows1.append([i, p.sku, p.descripcion, p.stock_referencial,
                                  p.predespacho_total, p.disponible,
                                  d.nombre, c.nombre, d.cantidad] + wh_vals)
                    i += 1
            else:
                rows1.append([i, p.sku, p.descripcion, p.stock_referencial,
                              p.predespacho_total, p.disponible,
                              "", c.nombre, c.total] + wh_vals)
                i += 1

    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Con Color"
    _write_sheet(ws1, "Con Color", headers1, rows1)

    # ── Sheet 2: Sin Color ─────────────────────────────────────────────
    headers2 = ["#", "SKU", "Descripción", "Stock", "Predesp", "Disponible"]
    for w in warehouses:
        headers2.append(w)

    rows2 = []
    for j, p in enumerate(sorted(productos, key=lambda x: x.sku), start=1):
        if _has_real_color(p):
            continue
        rows2.append([j, p.sku, p.descripcion, p.stock_referencial,
                       p.predespacho_total, p.disponible] + _warehouse_vals(p.sku))

    ws2 = wb.create_sheet("Sin Color")
    _write_sheet(ws2, "Sin Color", headers2, rows2)

    wb.save(path)